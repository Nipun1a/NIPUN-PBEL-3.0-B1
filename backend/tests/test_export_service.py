"""
Unit tests for backend/services/export_service.py.

Tests:
  - build_attendance_workbook: returns bytes (valid .xlsx)
  - build_attendance_workbook: header row is correct
  - build_attendance_workbook: empty input returns headers-only
  - build_attendance_workbook: row count matches records
  - build_attendance_workbook: data values match records
  - build_students_workbook: returns bytes (valid .xlsx)
  - build_students_workbook: header row is correct
  - build_students_workbook: empty input returns headers-only
  - build_students_workbook: row count matches students
"""
import sys
import os
from io import BytesIO

import pytest
import openpyxl

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from backend.services.export_service import build_attendance_workbook, build_students_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_workbook(xlsx_bytes: bytes) -> openpyxl.Workbook:
    """Parse raw bytes as an openpyxl workbook."""
    return openpyxl.load_workbook(BytesIO(xlsx_bytes))


def _get_sheet_rows(xlsx_bytes: bytes, sheet_title: str = None):
    """Return list of tuples for each row in the first (or named) sheet."""
    wb = _parse_workbook(xlsx_bytes)
    ws = wb[sheet_title] if sheet_title else wb.active
    return list(ws.iter_rows(values_only=True))


SAMPLE_ATTENDANCE_RECORDS = [
    {
        "roll_number": "101",
        "name": "Alice",
        "department": "CS",
        "date": "2024-01-15",
        "time": "09:05:00",
        "confidence_score": 0.87,
        "status": "Present",
        "marked_by": "face_recognition",
    },
    {
        "roll_number": "102",
        "name": "Bob",
        "department": "Math",
        "date": "2024-01-15",
        "time": "09:10:00",
        "confidence_score": 0.75,
        "status": "Present",
        "marked_by": "manual",
    },
]

SAMPLE_STUDENT_RECORDS = [
    {
        "roll_number": "101",
        "name": "Alice",
        "department": "CS",
        "email": "alice@example.com",
        "phone": "1234567890",
        "created_at": "2024-01-01T00:00:00Z",
    },
    {
        "roll_number": "102",
        "name": "Bob",
        "department": "Math",
        "email": "bob@example.com",
        "phone": "",
        "created_at": "2024-01-02T00:00:00Z",
    },
]


# ---------------------------------------------------------------------------
# build_attendance_workbook
# ---------------------------------------------------------------------------

class TestBuildAttendanceWorkbook:
    def test_returns_bytes(self):
        result = build_attendance_workbook([])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_format(self):
        result = build_attendance_workbook([])
        # Should parse without raising
        wb = _parse_workbook(result)
        assert wb is not None

    def test_sheet_title_is_attendance(self):
        result = build_attendance_workbook([])
        wb = _parse_workbook(result)
        assert "Attendance" in wb.sheetnames

    def test_header_row_correct(self):
        result = build_attendance_workbook([])
        rows = _get_sheet_rows(result, "Attendance")
        assert rows[0] == (
            "Roll Number", "Name", "Department", "Date",
            "Time", "Confidence Score", "Status", "Marked By",
        )

    def test_empty_records_returns_headers_only(self):
        result = build_attendance_workbook([])
        rows = _get_sheet_rows(result, "Attendance")
        assert len(rows) == 1  # headers only

    def test_row_count_matches_records(self):
        result = build_attendance_workbook(SAMPLE_ATTENDANCE_RECORDS)
        rows = _get_sheet_rows(result, "Attendance")
        # 1 header + 2 data rows
        assert len(rows) == 3

    def test_first_data_row_values(self):
        result = build_attendance_workbook(SAMPLE_ATTENDANCE_RECORDS)
        rows = _get_sheet_rows(result, "Attendance")
        data_row = rows[1]
        assert data_row[0] == "101"       # roll_number
        assert data_row[1] == "Alice"     # name
        assert data_row[2] == "CS"        # department
        assert data_row[3] == "2024-01-15"  # date
        assert data_row[4] == "09:05:00"  # time
        assert data_row[5] == pytest.approx(0.87)  # confidence_score
        assert data_row[6] == "Present"   # status
        assert data_row[7] == "face_recognition"  # marked_by

    def test_second_data_row_values(self):
        result = build_attendance_workbook(SAMPLE_ATTENDANCE_RECORDS)
        rows = _get_sheet_rows(result, "Attendance")
        data_row = rows[2]
        assert data_row[0] == "102"
        assert data_row[7] == "manual"

    def test_missing_department_defaults_to_empty(self):
        """When 'department' key is absent, the cell should be empty (None or "")."""
        records = [{"roll_number": "101", "name": "Alice",
                    "date": "2024-01-15", "time": "09:00:00",
                    "confidence_score": 0.9, "status": "Present",
                    "marked_by": "face_recognition"}]
        result = build_attendance_workbook(records)
        rows = _get_sheet_rows(result, "Attendance")
        # openpyxl reads empty string cells back as None
        assert rows[1][2] in ("", None)  # department is empty / missing


# ---------------------------------------------------------------------------
# build_students_workbook
# ---------------------------------------------------------------------------

class TestBuildStudentsWorkbook:
    def test_returns_bytes(self):
        result = build_students_workbook([])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_format(self):
        result = build_students_workbook([])
        wb = _parse_workbook(result)
        assert wb is not None

    def test_sheet_title_is_students(self):
        result = build_students_workbook([])
        wb = _parse_workbook(result)
        assert "Students" in wb.sheetnames

    def test_header_row_correct(self):
        result = build_students_workbook([])
        rows = _get_sheet_rows(result, "Students")
        assert rows[0] == (
            "Roll Number", "Name", "Department", "Email", "Phone", "Created At"
        )

    def test_empty_records_returns_headers_only(self):
        result = build_students_workbook([])
        rows = _get_sheet_rows(result, "Students")
        assert len(rows) == 1

    def test_row_count_matches_students(self):
        result = build_students_workbook(SAMPLE_STUDENT_RECORDS)
        rows = _get_sheet_rows(result, "Students")
        assert len(rows) == 3  # 1 header + 2 data rows

    def test_first_data_row_values(self):
        result = build_students_workbook(SAMPLE_STUDENT_RECORDS)
        rows = _get_sheet_rows(result, "Students")
        data_row = rows[1]
        assert data_row[0] == "101"
        assert data_row[1] == "Alice"
        assert data_row[2] == "CS"
        assert data_row[3] == "alice@example.com"
        assert data_row[4] == "1234567890"
        assert data_row[5] == "2024-01-01T00:00:00Z"
